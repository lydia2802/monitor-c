"""
Advanced Export & Reporting
Professional reports dengan charts, graphs, PDF generation, Excel dengan formulas.
"""

import os
import sys
import json
from datetime import datetime

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from pegasus.utils.helpers import print_colored, ensure_export_dir, format_timestamp

# Try to import optional dependencies
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import matplotlib
    matplotlib.use('Agg')  # Use non-interactive backend
    import matplotlib.pyplot as plt
    import io
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False


class AdvancedExporter:
    """Advanced exporter for professional reports"""
    
    def __init__(self):
        self.export_dir = "exports"
        ensure_export_dir()
    
    def export_to_pdf(self, search_history, filename="report.pdf"):
        """Export comprehensive PDF report"""
        if not REPORTLAB_AVAILABLE:
            print_colored("[!] ReportLab not installed. Cannot generate PDF.", "ERROR")
            print_colored("[i] Install with: pip install reportlab", "INFO")
            return None
        
        if not search_history:
            print_colored("[!] No search history available", "WARNING")
            return None
        
        filepath = os.path.join(self.export_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("<b>PEGASUS LACAK NOMOR</b><br/>Comprehensive Search Report", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 0.3*inch))
        
        # Generation info
        gen_info = Paragraph(f"Generated: {format_timestamp()}", styles['Normal'])
        story.append(gen_info)
        story.append(Spacer(1, 0.2*inch))
        
        # Summary stats
        stats = self._calculate_statistics(search_history)
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Searches', str(stats['total_searches'])],
            ['Phone Numbers', str(stats['phone_numbers'])],
            ['NIK Searches', str(stats['niks'])],
            ['Date Range', f"{stats['first_search']} to {stats['last_search']}"],
            ['Unique Targets', str(stats['unique_targets'])],
            ['Top City', stats.get('top_city', 'N/A')],
            ['Top Operator', stats.get('top_operator', 'N/A')]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E7E6E6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 0.5*inch))
        
        # Search details table
        story.append(Paragraph("<b>Recent Search Details</b>", styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))
        
        details_data = [['#', 'Target', 'Name', 'City', 'Operator', 'Date']]
        for i, entry in enumerate(search_history[:20], 1):  # Last 20
            result = entry.get('result', {})
            details_data.append([
                str(i),
                entry.get('target', '')[:15],
                result.get('Nama', 'N/A')[:20],
                result.get('Kota/Town', 'N/A')[:15],
                result.get('Operator', 'N/A')[:10],
                entry.get('timestamp', '')[:10]
            ])
        
        details_table = Table(details_data, repeatRows=1)
        details_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')])
        ]))
        
        story.append(details_table)
        
        # Build PDF
        doc.build(story)
        print_colored(f"[✓] PDF report generated: {filepath}", "SUCCESS")
        return filepath
    
    def export_to_excel_advanced(self, search_history, filename="report.xlsx"):
        """Export to Excel with charts and formulas"""
        if not OPENPYXL_AVAILABLE:
            print_colored("[!] openpyxl not installed. Cannot generate Excel.", "ERROR")
            print_colored("[i] Install with: pip install openpyxl", "INFO")
            return None
        
        if not search_history:
            print_colored("[!] No search history available", "WARNING")
            return None
        
        filepath = os.path.join(self.export_dir, filename)
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws_summary['A1'] = "PEGASUS LACAK NOMOR - COMPREHENSIVE REPORT"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:B1')
        
        ws_summary['A2'] = f"Generated: {format_timestamp()}"
        ws_summary.merge_cells('A2:B2')
        
        # Summary data
        stats = self._calculate_statistics(search_history)
        summary_rows = [
            ['Metric', 'Value'],
            ['Total Searches', stats['total_searches']],
            ['Phone Numbers', stats['phone_numbers']],
            ['NIK Searches', stats['niks']],
            ['Unique Targets', stats['unique_targets']],
            ['First Search', stats['first_search']],
            ['Last Search', stats['last_search']],
            ['Top City', stats.get('top_city', 'N/A')],
            ['Top Operator', stats.get('top_operator', 'N/A')]
        ]
        
        for row_idx, row_data in enumerate(summary_rows, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 4:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 30
        
        # Sheet 2: Detailed Data
        ws_details = wb.create_sheet("Detailed Data")
        
        # Headers
        detail_headers = ['#', 'Target', 'Name', 'Gender', 'City', 'Province', 'Operator', 'Timestamp']
        for col, header in enumerate(detail_headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Data
        for row_idx, entry in enumerate(search_history, 2):
            result = entry.get('result', {})
            ws_details.cell(row=row_idx, column=1, value=row_idx-1)
            ws_details.cell(row=row_idx, column=2, value=entry.get('target', ''))
            ws_details.cell(row=row_idx, column=3, value=result.get('Nama', ''))
            ws_details.cell(row=row_idx, column=4, value=result.get('Jenis Kelamin', ''))
            ws_details.cell(row=row_idx, column=5, value=result.get('Kota/Town', ''))
            ws_details.cell(row=row_idx, column=6, value=result.get('Provinsi', ''))
            ws_details.cell(row=row_idx, column=7, value=result.get('Operator', ''))
            ws_details.cell(row=row_idx, column=8, value=entry.get('timestamp', ''))
            
            # Add border to all cells
            for col in range(1, 9):
                ws_details.cell(row=row_idx, column=col).border = border
        
        # Adjust column widths
        for col_letter, width in [('A', 5), ('B', 20), ('C', 25), ('D', 10), 
                                   ('E', 20), ('F', 20), ('G', 15), ('H', 20)]:
            ws_details.column_dimensions[col_letter].width = width
        
        # Sheet 3: Analytics
        ws_charts = wb.create_sheet("Analytics")
        
        # Operator distribution
        ws_charts['A1'] = "Operator Distribution"
        ws_charts['A1'].font = Font(bold=True, size=12)
        
        ws_charts['A3'] = "Operator"
        ws_charts['B3'] = "Count"
        for col in ['A', 'B']:
            ws_charts[f'{col}3'].fill = header_fill
            ws_charts[f'{col}3'].font = header_font
        
        # Count operators
        operator_counts = self._get_operator_counts(search_history)
        for row_idx, (operator, count) in enumerate(operator_counts.items(), 4):
            ws_charts.cell(row=row_idx, column=1, value=operator)
            ws_charts.cell(row=row_idx, column=2, value=count)
        
        # Add bar chart for operators
        if operator_counts and len(operator_counts) > 0:
            chart = BarChart()
            chart.title = "Searches by Operator"
            chart.x_axis.title = "Operator"
            chart.y_axis.title = "Count"
            
            data = Reference(ws_charts, min_col=2, min_row=3, max_row=len(operator_counts)+3)
            cats = Reference(ws_charts, min_col=1, min_row=4, max_row=len(operator_counts)+3)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws_charts.add_chart(chart, "D3")
        
        # City distribution
        city_start_row = len(operator_counts) + 6
        ws_charts[f'A{city_start_row}'] = "City Distribution"
        ws_charts[f'A{city_start_row}'].font = Font(bold=True, size=12)
        
        ws_charts[f'A{city_start_row+2}'] = "City"
        ws_charts[f'B{city_start_row+2}'] = "Count"
        for col in ['A', 'B']:
            ws_charts[f'{col}{city_start_row+2}'].fill = header_fill
            ws_charts[f'{col}{city_start_row+2}'].font = header_font
        
        # Count cities
        city_counts = self._get_city_counts(search_history)
        for row_idx, (city, count) in enumerate(city_counts.items(), city_start_row+3):
            ws_charts.cell(row=row_idx, column=1, value=city)
            ws_charts.cell(row=row_idx, column=2, value=count)
        
        # Save workbook
        wb.save(filepath)
        print_colored(f"[✓] Excel report generated: {filepath}", "SUCCESS")
        return filepath
    
    def export_to_html_report(self, search_history, filename="report.html"):
        """Export to HTML report with interactive elements"""
        if not search_history:
            print_colored("[!] No search history available", "WARNING")
            return None
        
        filepath = os.path.join(self.export_dir, filename)
        stats = self._calculate_statistics(search_history)
        
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Pegasus Lacak Nomor - Report</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #366092;
            border-bottom: 3px solid #366092;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #366092;
            margin-top: 30px;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .stat-card {{
            background: linear-gradient(135deg, #366092, #4a7ab8);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}
        .stat-value {{
            font-size: 32px;
            font-weight: bold;
        }}
        .stat-label {{
            font-size: 14px;
            opacity: 0.9;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background-color: #366092;
            color: white;
            padding: 12px;
            text-align: left;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        tr:hover {{
            background-color: #f0f0f0;
        }}
        .footer {{
            margin-top: 40px;
            text-align: center;
            color: #666;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>PEGASUS LACAK NOMOR - COMPREHENSIVE REPORT</h1>
        <p><strong>Generated:</strong> {format_timestamp()}</p>
        
        <h2>Summary Statistics</h2>
        <div class="stats-grid">
            <div class="stat-card">
                <div class="stat-value">{stats['total_searches']}</div>
                <div class="stat-label">Total Searches</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats['phone_numbers']}</div>
                <div class="stat-label">Phone Numbers</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats['niks']}</div>
                <div class="stat-label">NIK Searches</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats['unique_targets']}</div>
                <div class="stat-label">Unique Targets</div>
            </div>
        </div>
        
        <h2>Recent Searches</h2>
        <table>
            <thead>
                <tr>
                    <th>#</th>
                    <th>Target</th>
                    <th>Name</th>
                    <th>City</th>
                    <th>Operator</th>
                    <th>Timestamp</th>
                </tr>
            </thead>
            <tbody>
"""
        
        # Add rows
        for i, entry in enumerate(search_history[:50], 1):
            result = entry.get('result', {})
            html_content += f"""
                <tr>
                    <td>{i}</td>
                    <td>{entry.get('target', '')}</td>
                    <td>{result.get('Nama', 'N/A')}</td>
                    <td>{result.get('Kota/Town', 'N/A')}</td>
                    <td>{result.get('Operator', 'N/A')}</td>
                    <td>{entry.get('timestamp', '')}</td>
                </tr>
"""
        
        html_content += """
            </tbody>
        </table>
        
        <div class="footer">
            <p>Generated by Pegasus Lacak Nomor v3.0</p>
        </div>
    </div>
</body>
</html>
"""
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print_colored(f"[✓] HTML report generated: {filepath}", "SUCCESS")
        return filepath
    
    def _calculate_statistics(self, history):
        """Calculate statistics from search history"""
        if not history:
            return {
                'total_searches': 0,
                'phone_numbers': 0,
                'niks': 0,
                'unique_targets': 0,
                'first_search': 'N/A',
                'last_search': 'N/A',
                'top_city': 'N/A',
                'top_operator': 'N/A'
            }
        
        total_searches = len(history)
        targets = [h.get('target', '') for h in history]
        phone_numbers = sum(1 for t in targets if t.startswith('08'))
        niks = total_searches - phone_numbers
        unique_targets = len(set(targets))
        
        # Get top city
        city_counts = self._get_city_counts(history)
        top_city = max(city_counts.items(), key=lambda x: x[1])[0] if city_counts else 'N/A'
        
        # Get top operator
        operator_counts = self._get_operator_counts(history)
        top_operator = max(operator_counts.items(), key=lambda x: x[1])[0] if operator_counts else 'N/A'
        
        return {
            'total_searches': total_searches,
            'phone_numbers': phone_numbers,
            'niks': niks,
            'unique_targets': unique_targets,
            'first_search': history[-1].get('timestamp', 'N/A')[:10] if history else 'N/A',
            'last_search': history[0].get('timestamp', 'N/A')[:10] if history else 'N/A',
            'top_city': top_city,
            'top_operator': top_operator
        }
    
    def _get_operator_counts(self, history):
        """Get counts by operator"""
        counts = {}
        for entry in history:
            operator = entry.get('result', {}).get('Operator')
            if operator:
                counts[operator] = counts.get(operator, 0) + 1
        return counts
    
    def _get_city_counts(self, history):
        """Get counts by city"""
        counts = {}
        for entry in history:
            city = entry.get('result', {}).get('Kota/Town')
            if city:
                counts[city] = counts.get(city, 0) + 1
        return counts


def generate_professional_report(search_history, format='all'):
    """Generate professional report in specified format"""
    exporter = AdvancedExporter()
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    files = []
    
    if format in ['pdf', 'all']:
        pdf_file = exporter.export_to_pdf(search_history, f"report_{timestamp}.pdf")
        if pdf_file:
            files.append(pdf_file)
    
    if format in ['excel', 'xlsx', 'all']:
        excel_file = exporter.export_to_excel_advanced(search_history, f"report_{timestamp}.xlsx")
        if excel_file:
            files.append(excel_file)
    
    if format in ['html', 'all']:
        html_file = exporter.export_to_html_report(search_history, f"report_{timestamp}.html")
        if html_file:
            files.append(html_file)
    
    return files
